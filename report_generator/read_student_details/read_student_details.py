__author__ = "Tafadzwa Brian Motsi"

import openpyxl
import warnings

warnings.filterwarnings("ignore", category=DeprecationWarning)


class StudentDetails:

    def student_details(self, student_details_filename):

        try:
            work_book = openpyxl.load_workbook(student_details_filename)
            sheets = work_book.sheetnames

            index = 0
            student_tests_marks = 0
            count = 0

            list_index = 0
            students_info_dict = {}

            while index < len(sheets):
                work_sheet = work_book.get_sheet_by_name(sheets[index])

                for row_index in range(2, work_sheet.max_row + 1):

                    key = work_sheet.cell(row=row_index, column=1).value
                    if key not in students_info_dict:
                        students_info_dict[key] = []
                        students_info_dict[key].append([])
                        students_info_dict[key][list_index].append(work_sheet.title)

                    else:
                        students_info_dict[key].append([])
                        students_info_dict[key][list_index].append(work_sheet.title)

                    for column_index in range(2, work_sheet.max_column):
                        student_tests_marks = student_tests_marks + work_sheet.cell(row=row_index, column=column_index)\
                            .value
                        count = count + 1

                    tests_average = int(round(student_tests_marks / count))
                    student_exam_mark = int(work_sheet.cell(row=row_index, column=work_sheet.max_column).value)

                    students_info_dict[key][list_index].append(tests_average)
                    students_info_dict[key][list_index].append(student_exam_mark)
                    students_info_dict[key][list_index].append(self.round_up((tests_average + student_exam_mark)*0.5))
                    student_tests_marks = 0
                    count = 0

                list_index = list_index + 1

                index = index + 1

        except IOError:
            print("could not read file: ", student_details_filename)

        return [students_info_dict, len(sheets)]

    @staticmethod
    def round_up(number):
        if number == int(number):
            return int(number)
        else:
            return int(number+1)
